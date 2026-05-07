#!/usr/bin/env python3
"""
Score the 73 business-rewritten revise instances using gpt-5.4.

The teacher's rubric (compressed by us into 5 explicit dimensions):

D1. Business Narrative Style (1-5)
    Does the revised business_requirement.md sound like something a real
    researcher / planner / manager would *say* in plain business language,
    matching the voice of original_workspace.docs.business_requirement.md
    (the "build" style)? 5 = indistinguishable in voice from original; 1 =
    reads like an instruction list / docstring.

D2. No Meta-Guidance Leakage (1-5)
    Does the text AVOID directly pointing at the model? Phrases such as
    "add a constraint ...", "modify the objective ...", "the model should
    enforce ...", "in addition to the original problem ...", "stating
    that ..." should NOT appear. 5 = none at all; 1 = it is essentially a
    list of formal modeling instructions.

D3. Numerical Specificity (1-5)
    Are penalty coefficients, thresholds, capacity bounds, ratios, time
    windows etc. all stated explicitly with concrete values or named
    parameters (e.g. `contract_nurse_pay`, `max_total_contract_starts`)?
    Vague phrasing ("penalise overuse", "should be small") gets low scores.
    5 = every quantitative element is specific; 1 = mostly vague.

D4. Business Motivation Plausibility (1-5)
    Is the revision wrapped inside a credible business motivation
    (regulation / customer pressure / safety / supplier change / contract /
    ESG / maintenance / market shift / strategic directive ...)? Is the
    motivation coherent with the underlying problem? 5 = rich, plausible,
    domain-appropriate motivation; 1 = no motivation, just a rule.

D5. Inference Difficulty - Build-Style Alignment (1-5)
    Does it require the solving LLM to TRANSLATE business prose into a
    constraint / objective change (not just lift a sentence verbatim)? Is
    its style aligned with `original_workspace.docs.business_requirement.md`
    so a model genuinely has to reason from business language to math?
    5 = forces real translation, indistinguishable in difficulty from the
    original build prompt; 1 = trivial; modeller can copy-paste.

The judge gpt-5.4 returns strict JSON with the five integer scores and a
short rationale (<= 60 words) per file. We then write an Excel:

    score_business_revise.xlsx
        Sheet "Scores":  num | file | D1 | D2 | D3 | D4 | D5 | total |
                         rationale
        Sheet "Stats":   per-dimension mean / median / min / max
        Sheet "Texts":   the original / old-revise / new-revise text per file
                         so the user can spot-check
"""
from __future__ import annotations
import os, sys, json, time, argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
from openai import OpenAI

API_KEY  = os.environ.get("OR_CLAW_API_KEY", "REDACTED_SET_ENV_VAR")
BASE_URL = "https://api.shubiaobiao.cn/v1"
MODEL    = "gpt-5.4"

BASE  = "."
SRC   = os.path.join(BASE, "IndustryOR_Revise_100")          # for "old revise"
NEW   = os.path.join(BASE, "IndustryOR_Revise_100_business") # for "new revise"

# OUT_X / LOGD are chosen in main() based on --out-tag so that the second-round
# retry scoring does not overwrite the first-round full-sweep artefacts.
OUT_X = os.path.join(BASE, "score_business_revise_full.xlsx")
LOGD  = os.path.join(BASE, "logs_score_business_full")

# Full 100-problem sweep (post-rename).
VERIFIED = list(range(1, 101))

DIMS = ["D1_business_style","D2_no_meta_guidance","D3_numerical_specificity",
        "D4_motivation_plausibility","D5_inference_difficulty"]

SYS_PROMPT = """You are a strict reviewer of operations-research revise prompts.

Your task: given (a) the ORIGINAL build-style business_requirement.md
(considered the gold style anchor), (b) the OLD mechanically-written revised
business_requirement.md, and (c) the NEW business-narrative revised
business_requirement.md, score ONLY the NEW one along five dimensions, on a
1-5 integer scale. Higher = better.

Dimensions
----------
D1 Business Narrative Style:
   Does NEW sound like real business prose from a researcher/planner/manager,
   matching the voice of ORIGINAL? 5 = indistinguishable in voice; 1 = reads
   like an instruction list / docstring.

D2 No Meta-Guidance Leakage:
   Does NEW AVOID phrases that directly name the model, such as: "add a
   constraint", "modify the objective", "a new constraint states that",
   "the model should enforce", "in addition to the original problem", "the
   formulation must now"? 5 = none at all; 1 = essentially a list of formal
   modeling instructions.

D3 Numerical Specificity:
   Are penalty coefficients / thresholds / capacities / ratios / time
   windows stated with concrete numeric values or named parameters (e.g.
   `contract_nurse_pay`, `max_total_contract_starts`)? Vague wording such
   as "penalise overuse", "small enough", "reasonable cost" should be
   penalised. 5 = every quantitative element is specific; 1 = mostly vague.

D4 Business Motivation Plausibility:
   Is the revision wrapped in a credible, domain-appropriate business
   motivation (regulation, contract, ESG, safety, supplier change,
   maintenance, market shift, strategic directive, etc.)? 5 = rich and
   plausible; 1 = no motivation, only a rule.

D5 Inference Difficulty (Build-Style Alignment):
   Does NEW force the solving LLM to TRANSLATE business prose into a
   constraint/objective change (rather than copy a sentence verbatim)? Is
   its writing style aligned with ORIGINAL such that a model has to reason
   from business language to math? 5 = forces real translation,
   indistinguishable in difficulty from ORIGINAL; 1 = trivial.

Output format
-------------
Return STRICTLY a JSON object with these EXACT keys (no extra keys, no
prose around it):

{
  "D1_business_style":         <int 1-5>,
  "D2_no_meta_guidance":       <int 1-5>,
  "D3_numerical_specificity":  <int 1-5>,
  "D4_motivation_plausibility":<int 1-5>,
  "D5_inference_difficulty":   <int 1-5>,
  "rationale":                 "<= 60 words explaining the lowest scoring dim>"
}

Be a strict, numerically-careful judge. Use 5 only for genuinely excellent
cases.
"""

USER_TMPL = """[ORIGINAL build-style business_requirement.md]
{original}

[OLD revised business_requirement.md (mechanical, for reference of meaning)]
{old_revise}

[NEW revised business_requirement.md (the one you must score)]
{new_revise}

Now score the NEW revised text. Return ONLY the JSON described in the system message.
"""


def _find(num: int, root: str) -> str | None:
    # Supports three naming schemes after the rename:
    #   IndustryOR_<N>_revise.json            (new SRC dir)
    #   IndustryOR_<N>_revise_business.json   (new NEW dir)
    #   IndustryOR_<N>_revise_<k>.json        (legacy, pre-rename)
    exact_names = (
        f"IndustryOR_{num}_revise.json",
        f"IndustryOR_{num}_revise_business.json",
    )
    for en in exact_names:
        p = os.path.join(root, en)
        if os.path.isfile(p):
            return p
    prefix = f"IndustryOR_{num}_revise_"
    for fn in os.listdir(root):
        if fn.startswith(prefix) and fn.endswith(".json"):
            return os.path.join(root, fn)
    return None


def _strip_fences(s: str) -> str:
    s = s.strip()
    if s.startswith("```"):
        lines = s.splitlines()
        lines = lines[1:]
        if lines and lines[-1].strip().startswith("```"):
            lines = lines[:-1]
        s = "\n".join(lines).strip()
    return s


def _parse_scores(raw: str) -> dict:
    txt = _strip_fences(raw)
    # find first '{' and last '}'
    i = txt.find("{"); j = txt.rfind("}")
    if i == -1 or j == -1:
        raise ValueError(f"no JSON object: {raw[:200]}")
    obj = json.loads(txt[i:j+1])
    for d in DIMS:
        if d not in obj:
            raise ValueError(f"missing key {d}: {obj}")
        v = obj[d]
        if not isinstance(v, (int, float)) or v < 1 or v > 5:
            raise ValueError(f"score out of range for {d}: {v}")
        obj[d] = int(round(v))
    obj.setdefault("rationale", "")
    return obj


def _score_one(num: int, tries: int = 3) -> dict:
    src_path = _find(num, SRC)
    new_path = _find(num, NEW)
    if not src_path or not new_path:
        return {"num": num, "ok": False, "err": "json missing"}
    s = json.load(open(src_path))
    n = json.load(open(new_path))
    original   = s["original_workspace"]["docs"]["business_requirement.md"]
    old_revise = s["revised_workspace"]["docs"]["business_requirement.md"]
    new_revise = n["revised_workspace"]["docs"]["business_requirement.md"]
    file_name  = os.path.basename(new_path)

    user_msg = USER_TMPL.format(original=original, old_revise=old_revise,
                                new_revise=new_revise)
    client = OpenAI(api_key=API_KEY, base_url=BASE_URL, timeout=120)
    last_err = ""
    for i in range(tries):
        try:
            r = client.chat.completions.create(
                model=MODEL,
                messages=[
                    {"role": "system", "content": SYS_PROMPT},
                    {"role": "user",   "content": user_msg},
                ],
                temperature=0.0,
                max_tokens=600,
            )
            raw = r.choices[0].message.content or ""
            obj = _parse_scores(raw)
            return {"num": num, "file": file_name, "ok": True,
                    "original": original, "old_revise": old_revise,
                    "new_revise": new_revise, **obj}
        except Exception as e:
            last_err = str(e)
            time.sleep(2 + 3 * i)
    return {"num": num, "file": file_name, "ok": False,
            "err": f"llm_fail: {last_err}",
            "original": original, "old_revise": old_revise,
            "new_revise": new_revise}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--concurrency", type=int, default=6)
    ap.add_argument("--instances", type=str, default="")
    ap.add_argument("--out-tag", type=str, default="full",
                    help="suffix for xlsx/log dir: writes "
                         "score_business_revise_<tag>.xlsx and "
                         "logs_score_business_<tag>/")
    args = ap.parse_args()

    global OUT_X, LOGD
    OUT_X = os.path.join(BASE, f"score_business_revise_{args.out_tag}.xlsx")
    LOGD  = os.path.join(BASE, f"logs_score_business_{args.out_tag}")
    os.makedirs(LOGD, exist_ok=True)
    try: sys.stdout.reconfigure(line_buffering=True)
    except Exception: pass

    nums = ([int(x) for x in args.instances.split(",") if x.strip().isdigit()]
            if args.instances else VERIFIED)
    total = len(nums)
    print(f"Scoring {total} instances with model={MODEL}  concurrency={args.concurrency}",
          flush=True)

    results = []
    done = 0
    with ThreadPoolExecutor(max_workers=args.concurrency) as pool:
        futs = {pool.submit(_score_one, n): n for n in nums}
        for fut in as_completed(futs):
            num = futs[fut]
            try:
                res = fut.result()
            except Exception as e:
                res = {"num": num, "ok": False, "err": f"exc:{e}"}
            done += 1
            results.append(res)
            if res.get("ok"):
                tot = sum(res[d] for d in DIMS)
                print(f"[{done}/{total}] #{num} OK  D1-D5={res['D1_business_style']},"
                      f"{res['D2_no_meta_guidance']},{res['D3_numerical_specificity']},"
                      f"{res['D4_motivation_plausibility']},{res['D5_inference_difficulty']}"
                      f"  total={tot}", flush=True)
            else:
                print(f"[{done}/{total}] #{num} FAIL  {res.get('err','')[:120]}",
                      flush=True)

    # write Excel
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Scores"
    headers = ["num","file","D1_business_style","D2_no_meta_guidance",
               "D3_numerical_specificity","D4_motivation_plausibility",
               "D5_inference_difficulty","total","rationale"]
    ws.append(headers)
    for c in ws[1]: c.font = Font(bold=True); c.alignment = Alignment(horizontal="center")

    results_sorted = sorted(results, key=lambda r: r.get("num", 0))
    for res in results_sorted:
        if res.get("ok"):
            tot = sum(res[d] for d in DIMS)
            ws.append([res["num"], res.get("file",""),
                       res["D1_business_style"], res["D2_no_meta_guidance"],
                       res["D3_numerical_specificity"],
                       res["D4_motivation_plausibility"],
                       res["D5_inference_difficulty"],
                       tot, res.get("rationale","")])
        else:
            ws.append([res["num"], res.get("file",""), "", "", "", "", "", "",
                       f"FAIL: {res.get('err','')}"])
            last = ws.max_row
            ws.cell(row=last, column=1).fill = PatternFill("solid", fgColor="FFC7CE")

    # color cells: 1-2 red, 3 yellow, 4 light green, 5 green
    fills = {1:"FFC7CE", 2:"FFC7CE", 3:"FFEB9C", 4:"C6EFCE", 5:"63BE7B"}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                            min_col=3, max_col=7):
        for cell in row:
            v = cell.value
            if isinstance(v, int) and 1 <= v <= 5:
                cell.fill = PatternFill("solid", fgColor=fills[v])
                cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['I'].width = 60

    # Stats sheet
    ws2 = wb.create_sheet("Stats")
    ws2.append(["dimension","mean","median","min","max","#5","#1-2"])
    for c in ws2[1]: c.font = Font(bold=True)
    ok = [r for r in results if r.get("ok")]
    import statistics as st
    for d in DIMS:
        vals = [r[d] for r in ok]
        if not vals:
            ws2.append([d, "", "", "", "", "", ""]); continue
        ws2.append([d, round(st.mean(vals),3), st.median(vals), min(vals),
                    max(vals), sum(1 for v in vals if v == 5),
                    sum(1 for v in vals if v <= 2)])
    if ok:
        totals = [sum(r[d] for d in DIMS) for r in ok]
        ws2.append([])
        ws2.append(["total (sum of 5 dims)", round(st.mean(totals),3),
                    st.median(totals), min(totals), max(totals), "", ""])
    ws2.column_dimensions['A'].width = 32

    # Texts sheet
    ws3 = wb.create_sheet("Texts")
    ws3.append(["num","file","ORIGINAL build md","OLD revise md","NEW revise md"])
    for c in ws3[1]: c.font = Font(bold=True)
    for res in results_sorted:
        ws3.append([res.get("num",""), res.get("file",""),
                    (res.get("original") or "")[:32000],
                    (res.get("old_revise") or "")[:32000],
                    (res.get("new_revise") or "")[:32000]])
    for col in "ABCDE":
        ws3.column_dimensions[col].width = 60
    for row in ws3.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(OUT_X)
    print(f"\n✅ Excel written: {OUT_X}")

    # also dump raw results JSON
    with open(os.path.join(LOGD, "results.json"), "w", encoding="utf-8") as f:
        json.dump(results_sorted, f, ensure_ascii=False, indent=2)

    # console summary
    if ok:
        print("\n===== Summary =====")
        for d in DIMS:
            vs = [r[d] for r in ok]
            print(f"  {d:34s} mean={st.mean(vs):.2f}  med={st.median(vs)}  "
                  f"min={min(vs)} max={max(vs)}  #5={sum(1 for v in vs if v==5)}  "
                  f"#<=2={sum(1 for v in vs if v<=2)}")
        totals = [sum(r[d] for d in DIMS) for r in ok]
        print(f"  TOTAL                              mean={st.mean(totals):.2f}  "
              f"med={st.median(totals)}  min={min(totals)} max={max(totals)}")
    fails = [r for r in results if not r.get("ok")]
    if fails:
        print(f"\nFailed: {len(fails)}")
        for r in fails:
            print(f"  #{r['num']}: {r.get('err','')[:120]}")


if __name__ == "__main__":
    main()
